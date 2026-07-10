"""Generate src/data/lksHierarchy.json and refresh lcmsUserMap.json.

Sources (project root):
- LKS_Firm_Hierarchy (1).xlsx — Summary + Complete Hierarchy sheets
- LKS_Firm_Hierarchy_employee_to_lcms_user_id_enhanced.csv — employee_code -> LCMS user id
- LKS_Firm_MindMap (1).html — optional validation of the 10 group heads

Run:
  py scripts/generateLksHierarchy.py
"""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "LKS_Firm_Hierarchy (1).xlsx"
LCMS_CSV = ROOT / "LKS_Firm_Hierarchy_employee_to_lcms_user_id_enhanced.csv"
MINDMAP_HTML = ROOT / "LKS_Firm_MindMap (1).html"
OUT_HIER = ROOT / "src/data/lksHierarchy.json"
OUT_LCMS = ROOT / "src/data/lcmsUserMap.json"
MANUAL_OVERRIDES = ROOT / "src/data/lcmsManualOverrides.json"

PARTNER_DESIGS = {"Executive Partner", "Senior Partner", "Partner", "Associate Partner"}


def normalize_name(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", "", s.lower())
    return s


def division_to_pillar(division: str | None) -> str:
    d = (division or "").lower()
    if "corporate" in d or "competition" in d:
        return "Corporate Law"
    if "direct tax" in d:
        return "Direct Tax"
    if "dispute" in d or "investigation" in d or "aiit" in d:
        return "Disputes & Investigations"
    if "international trade" in d or "customs" in d:
        return "International Trade & Customs"
    if "ipr" in d or "intellectual" in d:
        return "Intellectual Property"
    return "Indirect Tax & GST"


def division_to_subpractice(division: str | None, sub: str | None) -> str:
    pillar = division_to_pillar(division)
    s = (sub or "").lower()
    if pillar == "Corporate Law":
        if "m&a" in s or "merger" in s:
            return "Mergers & Acquisitions"
        return "General Corporate"
    if pillar == "Direct Tax":
        if "transfer" in s:
            return "Transfer Pricing"
        if "international" in s:
            return "International Tax"
        return "Direct Tax"
    if pillar == "Disputes & Investigations":
        if "arbitr" in s:
            return "Arbitration"
        if "insolv" in s:
            return "Insolvency"
        if "economic" in s:
            return "Economic Offences"
        return "Commercial Litigation"
    if pillar == "International Trade & Customs":
        if "wto" in s or "trade remed" in s or "trade consult" in s:
            return "International Trade and WTO"
        return "Customs"
    if pillar == "Intellectual Property":
        if "patent" in s:
            return "Patents"
        if "trademark" in s:
            return "Trademarks"
        if "litig" in s:
            return "IP Litigation"
        return "IP Litigation"
    if "consult" in s or "advis" in s:
        return "GST Consultancy"
    return "GST & Indirect Tax Litigation"


# Post-processing: firm structure overrides (mind-map / dashboard layout).
EXCLUDED_SUMMARY_PH_NAMES = {normalize_name("Malathi Lakshmikumaran")}
DEVINDER_BAGIA_CODE = 1679
DEVINDER_GROUP_HEAD = "Devinder Bagia"
MALATHI_GROUP_HEAD = "Malathi L"
CHARANYA_LAKSHMIKUMARAN_CODE = 1123


def is_intl_trade_team(team_mapping: str | None) -> bool:
    return (team_mapping or "").startswith("International Trade")


def should_remove_malathi_ip_team(person: dict) -> bool:
    """Drop Malathi as group head and her IP litigation subtree."""
    if person.get("employeeCode") == 90002:
        return True
    if (person.get("practiceHeadName") or "") != MALATHI_GROUP_HEAD:
        return False
    return (person.get("teamMapping") or "").startswith("IPR Litigation")


def apply_hierarchy_adjustments(people: list[dict]) -> list[dict]:
    """Promote Devinder Bagia's intl trade group; remove Malathi IP litigation."""
    people = [p for p in people if not should_remove_malathi_ip_team(p)]

    devinder = next((p for p in people if p["employeeCode"] == DEVINDER_BAGIA_CODE), None)
    if devinder:
        devinder["role"] = "practice_head"
        devinder["supervisorId"] = None
        devinder["practiceHeadName"] = DEVINDER_GROUP_HEAD
        devinder["teamMapping"] = DEVINDER_GROUP_HEAD
        devinder["pillar"] = division_to_pillar("International Trade")
        devinder["subPractice"] = division_to_subpractice("International Trade", "WTO")

    for p in people:
        if is_intl_trade_team(p.get("teamMapping")):
            p["practiceHeadName"] = DEVINDER_GROUP_HEAD
            p["pillar"] = division_to_pillar("International Trade")

    prashant = next((p for p in people if p["employeeCode"] == 1097), None)
    if prashant:
        prashant["practiceHeadName"] = "Charanya L"
        prashant["supervisorId"] = f"lks-{CHARANYA_LAKSHMIKUMARAN_CODE}"
        prashant["reportingAuthority"] = "Charanya Lakshmikumaran"

    return people


def parse_employee_code_from_rp(rp: str | None) -> int | None:
    m = re.search(r"\((\d+)\)", rp or "")
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def reconcile_group_supervisors(people: list[dict]) -> None:
    """Keep reporting chains inside each practice-head group for UI drill-down."""
    by_id = {p["id"]: p for p in people}
    by_code = {p["employeeCode"]: p for p in people}
    name_to_id: dict[str, str] = {}
    for p in people:
        name_to_id.setdefault(normalize_name(p["name"]), p["id"])

    ph_by_group: dict[str, str] = {}
    for p in people:
        if p["role"] == "practice_head":
            ph_by_group[p["practiceHeadName"]] = p["id"]

    def in_group(person_id: str | None, group: str) -> bool:
        if not person_id:
            return False
        person = by_id.get(person_id)
        return bool(person and person.get("practiceHeadName") == group)

    for p in people:
        if p["role"] == "practice_head":
            continue
        group = (p.get("practiceHeadName") or "").strip()
        ph_id = ph_by_group.get(group)
        if not ph_id:
            continue

        current = p.get("supervisorId")
        if in_group(current, group):
            continue

        sup_id: str | None = None
        ra_norm = normalize_name(p.get("reportingAuthority") or "")
        if ra_norm:
            cand = name_to_id.get(ra_norm)
            if in_group(cand, group):
                sup_id = cand

        if not sup_id:
            rp_code = parse_employee_code_from_rp(p.get("reportingPartner"))
            if rp_code and rp_code in by_code:
                cand_id = by_code[rp_code]["id"]
                if cand_id != p["id"] and in_group(cand_id, group):
                    sup_id = cand_id

        if not sup_id:
            ph = by_id[ph_id]
            if ra_norm in {normalize_name(ph["name"]), normalize_name(group)}:
                sup_id = ph_id

        p["supervisorId"] = sup_id or ph_id


def apply_sub_heads(people: list[dict], summary) -> None:
    """Tag Summary-sheet sub-heads (e.g. Manish Gaur under B L Narsimhan)."""
    name_to_person = {normalize_name(p["name"]): p for p in people}
    for row in summary.iter_rows(min_row=5, values_only=True):
        _sno, group_head, _full_name, _practice_area, sub_head = row[:5]
        if not sub_head or str(sub_head).strip() in {"", "-"}:
            continue
        sub_name = str(sub_head).strip()
        person = name_to_person.get(normalize_name(sub_name))
        if not person:
            continue
        person["isSubHead"] = True
        person["subHeadFor"] = (group_head or "").strip()


def load_lcms_by_employee_code() -> dict[int, dict]:
    """employee_code -> { lcmsUserId, matchMethod, lcmsDisplayName }"""
    out: dict[int, dict] = {}
    with LCMS_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            code_s = (row.get("employee_code") or "").strip()
            lcms_s = (row.get("lcms_user_id") or "").strip()
            if not code_s or not lcms_s:
                continue
            try:
                code = int(code_s)
                lcms_id = int(float(lcms_s))
            except ValueError:
                continue
            out[code] = {
                "lcmsUserId": lcms_id,
                "matchMethod": (row.get("match_method") or "csv").strip(),
                "lcmsDisplayName": (row.get("lcms_display_name") or "").strip(),
            }
    return out


def load_manual_lcms_overrides() -> dict[int, dict]:
    """employee_code -> manual LCMS id (group heads absent from employee CSV)."""
    if not MANUAL_OVERRIDES.exists():
        return {}
    raw = json.loads(MANUAL_OVERRIDES.read_text(encoding="utf-8"))
    out: dict[int, dict] = {}
    for code_s, meta in raw.items():
        try:
            code = int(code_s)
            lcms_id = int(meta["lcmsUserId"])
        except (KeyError, TypeError, ValueError):
            continue
        out[code] = {
            "lcmsUserId": lcms_id,
            "matchMethod": (meta.get("matchMethod") or "manual").strip(),
            "lcmsDisplayName": (meta.get("name") or "").strip(),
        }
    return out


def merge_lcms_sources() -> dict[int, dict]:
    merged = load_lcms_by_employee_code()
    for code, meta in load_manual_lcms_overrides().items():
        merged[code] = meta
    return merged


def load_mindmap_group_heads() -> list[str]:
    """Extract group-head full_name values from the mind-map HTML for sanity checks."""
    if not MINDMAP_HTML.exists():
        return []
    text = MINDMAP_HTML.read_text(encoding="utf-8")
    m = re.search(r"const data = (\{.*?\});", text, re.DOTALL)
    if not m:
        return []
    data = json.loads(m.group(1))
    return [
        str(child.get("full_name") or child.get("name") or "").strip()
        for child in data.get("children", [])
        if child.get("role") == "Group Head"
    ]


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")
    if not LCMS_CSV.exists():
        raise SystemExit(f"Missing LCMS CSV: {LCMS_CSV}")

    lcms_by_code = merge_lcms_sources()
    wb = load_workbook(XLSX, data_only=True)

    # --- summary: practice-head full names ----------------------------------
    summary = wb["Summary"]
    ph_full_names: set[str] = set()
    ph_charters: dict[str, str] = {}
    for row in summary.iter_rows(min_row=5, values_only=True):
        _sno, _group_head, full_name, practice_area, *_ = row
        if not full_name:
            continue
        name = full_name.strip()
        if normalize_name(name) in EXCLUDED_SUMMARY_PH_NAMES:
            continue
        ph_full_names.add(name)
        ph_charters[name] = (practice_area or "").strip()
    ph_norm = {normalize_name(n) for n in ph_full_names}

    mindmap_gh = load_mindmap_group_heads()
    if mindmap_gh:
        mm_norm = {normalize_name(n) for n in mindmap_gh}
        missing_in_xlsx = [n for n in mindmap_gh if normalize_name(n) not in ph_norm]
        missing_in_mindmap = [n for n in ph_full_names if normalize_name(n) not in mm_norm]
        if missing_in_xlsx:
            print(f"WARN mindmap group heads absent from Summary sheet: {missing_in_xlsx}")
        if missing_in_mindmap:
            print(f"WARN Summary PHs absent from mindmap: {missing_in_mindmap}")

    # --- complete hierarchy --------------------------------------------------
    sheet = wb["Complete Hierarchy"]
    rows = list(sheet.iter_rows(min_row=3, values_only=True))

    people: list[dict] = []
    seen_codes: set[int] = set()
    name_to_code: dict[str, int] = {}

    for r in rows:
        if not r:
            continue
        rp = r[10] if len(r) > 10 else None
        if not rp:
            continue
        m = re.match(r"^(.*?)\s*\((\d+)\)\s*$", str(rp))
        if m:
            name_to_code.setdefault(normalize_name(m.group(1)), int(m.group(2)))

    for r in rows:
        if not r or r[0] is None:
            continue
        (
            _sno,
            _group_head,
            _practice_area,
            _sub_head,
            emp_code,
            emp_name,
            designation,
            location,
            team_mapping,
            reporting_authority,
            reporting_partner,
            division,
        ) = r[:12]
        if emp_code is None or not emp_name:
            continue
        try:
            code = int(emp_code)
        except (TypeError, ValueError):
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)

        name = str(emp_name).strip()
        nname = normalize_name(name)
        designation_s = (designation or "").strip()
        if nname in ph_norm:
            role = "practice_head"
        elif designation_s in PARTNER_DESIGS:
            role = "partner"
        else:
            role = "associate"

        lcms = lcms_by_code.get(code)
        person: dict = {
            "id": f"lks-{code}",
            "employeeCode": code,
            "name": name,
            "designation": designation_s,
            "role": role,
            "office": (location or "New Delhi").strip(),
            "division": (division or "").strip(),
            "subDivision": (r[12] if len(r) > 12 else None) or "",
            "teamMapping": (team_mapping or "").strip(),
            "reportingAuthority": (reporting_authority or "").strip(),
            "reportingPartner": (reporting_partner or "").strip(),
            "practiceHeadName": (_group_head or "").strip(),
            "pillar": division_to_pillar(division),
            "subPractice": division_to_subpractice(division, r[12] if len(r) > 12 else None),
        }
        if lcms:
            person["lcmsUserId"] = lcms["lcmsUserId"]
        people.append(person)

    # Inject practice heads named in Summary but absent from Complete Hierarchy.
    synthetic_code = 90000
    present_norms = {normalize_name(p["name"]) for p in people}
    summary_ph_meta: dict[str, dict] = {}
    for row in summary.iter_rows(min_row=5, values_only=True):
        _sno, group_head, full_name, practice_area, *_ = row
        if full_name:
            summary_ph_meta[normalize_name(full_name)] = {
                "name": full_name.strip(),
                "groupHead": (group_head or "").strip(),
                "practiceArea": (practice_area or "").strip(),
            }
    for n_norm, meta in summary_ph_meta.items():
        if n_norm in EXCLUDED_SUMMARY_PH_NAMES:
            continue
        if n_norm in present_norms:
            continue
        code = name_to_code.get(n_norm)
        if code is None or code in seen_codes:
            synthetic_code += 1
            while synthetic_code in seen_codes:
                synthetic_code += 1
            code = synthetic_code
        seen_codes.add(code)
        person = {
            "id": f"lks-{code}",
            "employeeCode": code,
            "name": meta["name"],
            "designation": "Senior Partner",
            "role": "practice_head",
            "office": "New Delhi",
            "division": "",
            "subDivision": "",
            "teamMapping": meta["groupHead"],
            "reportingAuthority": "V. Lakshmikumaran",
            "reportingPartner": "V. Lakshmikumaran (1001)",
            "practiceHeadName": meta["groupHead"],
            "pillar": division_to_pillar(meta["practiceArea"]),
            "subPractice": division_to_subpractice(meta["practiceArea"], None),
        }
        lcms = lcms_by_code.get(code)
        if lcms:
            person["lcmsUserId"] = lcms["lcmsUserId"]
        people.append(person)

    # --- supervisorId by name match -----------------------------------------
    name_to_id: dict[str, str] = {}
    for p in people:
        name_to_id.setdefault(normalize_name(p["name"]), p["id"])

    for p in people:
        if p["role"] == "practice_head":
            p["supervisorId"] = None
            continue
        sup_norm = normalize_name(p["reportingAuthority"])
        sup_id = name_to_id.get(sup_norm)
        if not sup_id:
            rp = p["reportingPartner"]
            m = re.search(r"\((\d+)\)", rp)
            if m:
                rp_code = int(m.group(1))
                sup_id = f"lks-{rp_code}" if rp_code in seen_codes else None
            if not sup_id:
                rp_clean = re.sub(r"\s*\(\d+\)\s*$", "", rp).strip()
                sup_id = name_to_id.get(normalize_name(rp_clean))
        p["supervisorId"] = sup_id

    people = apply_hierarchy_adjustments(people)
    reconcile_group_supervisors(people)
    apply_sub_heads(people, summary)

    # Manual overrides for group heads not present in the employee CSV.
    for p in people:
        lcms = lcms_by_code.get(p["employeeCode"])
        if lcms:
            p["lcmsUserId"] = lcms["lcmsUserId"]

    matched = sum(1 for p in people if "lcmsUserId" in p)
    OUT_HIER.write_text(json.dumps(people, indent=2))

    new_map: dict[str, dict] = {}
    for p in people:
        if "lcmsUserId" not in p:
            continue
        code = p["employeeCode"]
        lcms_meta = lcms_by_code.get(code, {})
        new_map[p["id"]] = {
            "lcmsUserId": p["lcmsUserId"],
            "name": p["name"],
            "matchMethod": lcms_meta.get("matchMethod", "csv"),
        }

    OUT_LCMS.write_text(json.dumps(new_map, indent=2))

    role_counts = {"practice_head": 0, "partner": 0, "associate": 0}
    for p in people:
        role_counts[p["role"]] += 1
    no_sup = sum(1 for p in people if p["role"] != "practice_head" and not p.get("supervisorId"))
    unmatched_codes = sorted(
        p["employeeCode"] for p in people if "lcmsUserId" not in p
    )

    print(f"Wrote {len(people)} people -> {OUT_HIER.relative_to(ROOT)}")
    print(f"Roles: {role_counts}")
    print(f"LCMS matches: {matched} / {len(people)} -> {OUT_LCMS.relative_to(ROOT)}")
    print(f"People without supervisor (non-PH): {no_sup}")
    if unmatched_codes:
        print(f"Unmatched employee codes ({len(unmatched_codes)}): {unmatched_codes[:15]}{'...' if len(unmatched_codes) > 15 else ''}")


if __name__ == "__main__":
    main()
